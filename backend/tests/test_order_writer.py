"""Tests for the low-level Excel writer (excel/order_writer.py).

All tests build a small synthetic template (via pytest's tmp_path) that mirrors the
verified Hikma structure — none of them touch the real
backend/templates/Hikma orders.xlsx workbook.
"""

import hashlib
import zipfile

import openpyxl
import pytest
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from excel.order_writer import (
    OrderLine,
    TemplateUnavailableError,
    build_output_filename,
    generate_order_workbook,
    sanitize_filename_component,
)


def _build_template_workbook(tmp_path, filename="template.xlsx"):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = "اسم الزبون"
    worksheet["A1"].font = Font(bold=True, size=12, color="FFFF0000")
    worksheet["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Product name",
        " price (Drug Store)",
        " price (Pharmacy & Hospitals)",
        "Qty",
        "FQty",
        "Value ",
        "Value",
        "Note ",
    ]
    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=2, column=col, value=header)
        worksheet.cell(row=2, column=col).fill = PatternFill("solid", fgColor="FFBDD7EE")
        worksheet.cell(row=2, column=col).border = Border(bottom=Side(style="thin"))

    products = [
        ("Alpha Tablet 50MG", 1000, 1200),
        ("Beta Syrup", 2000, 2400),
        ("Gamma Injection", 500, 600),
    ]
    for index, (name, price_b, price_c) in enumerate(products):
        row = 3 + index
        worksheet.cell(row=row, column=1, value=name)
        worksheet.cell(row=row, column=2, value=price_b)
        worksheet.cell(row=row, column=3, value=price_c)
        worksheet.cell(row=row, column=6, value=f"=D{row}*B{row}")
        worksheet.cell(row=row, column=7, value=f"=D{row}*C{row}")
        for column in (2, 3, 6, 7):
            worksheet.cell(row=row, column=column).number_format = '#,##0 "IQD"'

    worksheet.cell(row=15, column=6, value="=SUM(F3:F14)")
    worksheet.cell(row=15, column=7, value="=SUM(G3:G14)")
    worksheet.cell(row=15, column=6).font = Font(bold=True)
    worksheet.cell(row=15, column=7).font = Font(bold=True)
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 30
    worksheet.row_dimensions[15].height = 22
    worksheet.print_area = "A1:H15"
    worksheet.page_setup.orientation = "portrait"

    for letter, width in [("A", 43.58), ("B", 14.39), ("C", 14.93), ("D", 6.46), ("E", 6.19), ("F", 16.01), ("G", 15.20)]:
        worksheet.column_dimensions[letter].width = width

    path = tmp_path / filename
    workbook.save(str(path))
    return path


def _file_hash(path) -> str:
    with open(path, "rb") as handle:
        return hashlib.sha256(handle.read()).hexdigest()


def _load(path):
    workbook = openpyxl.load_workbook(str(path))
    worksheet = workbook["Sheet1"]
    return workbook, worksheet


# --- source safety -----------------------------------------------------------------------


def test_source_workbook_is_never_modified(tmp_path):
    source = _build_template_workbook(tmp_path, "source.xlsx")
    before_hash = _file_hash(source)

    generate_order_workbook(
        order_title="Test Customer",
        order_lines=[OrderLine(row=3, quantity=5, free_quantity=1)],
        selected_price_type="pharmacy",
        output_path=tmp_path / "output.xlsx",
        source_path=source,
    )

    assert _file_hash(source) == before_hash


def test_source_workbook_hash_unchanged_after_generation(tmp_path):
    # Same intent as the previous test, kept as its own case to match the required list 1:1.
    source = _build_template_workbook(tmp_path, "source_hash.xlsx")
    before_hash = _file_hash(source)

    generate_order_workbook(
        order_title="Another Customer",
        order_lines=[OrderLine(row=4, quantity=2, free_quantity=0)],
        selected_price_type="drug_store",
        output_path=tmp_path / "output_hash.xlsx",
        source_path=source,
    )

    assert _file_hash(source) == before_hash


def test_output_is_saved_to_a_different_path(tmp_path):
    source = _build_template_workbook(tmp_path, "source2.xlsx")
    output_path = tmp_path / "generated" / "output2.xlsx"

    result = generate_order_workbook(
        order_title="Test Customer",
        order_lines=[OrderLine(row=3, quantity=1, free_quantity=0)],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    assert result.output_path == output_path
    assert output_path != source
    assert output_path.exists()


def test_missing_source_template_is_handled_safely(tmp_path):
    with pytest.raises(TemplateUnavailableError):
        generate_order_workbook(
            order_title="Test Customer",
            order_lines=[],
            selected_price_type="pharmacy",
            output_path=tmp_path / "output.xlsx",
            source_path=tmp_path / "does-not-exist.xlsx",
        )


# --- writing behavior ----------------------------------------------------------------------


def test_customer_title_written_to_correct_cell(tmp_path):
    source = _build_template_workbook(tmp_path, "source3.xlsx")
    output_path = tmp_path / "output3.xlsx"

    generate_order_workbook(
        order_title="صيدلية العين - النجف",
        order_lines=[],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    assert worksheet["A1"].value == "صيدلية العين - النجف"


def test_quantities_written_to_correct_product_rows(tmp_path):
    source = _build_template_workbook(tmp_path, "source4.xlsx")
    output_path = tmp_path / "output4.xlsx"

    generate_order_workbook(
        order_title="Test",
        order_lines=[
            OrderLine(row=3, quantity=5, free_quantity=0),
            OrderLine(row=5, quantity=12, free_quantity=0),
        ],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    assert worksheet["C3"].value == 5
    assert worksheet["C4"].value == 12
    assert worksheet.max_row == 5


def test_free_quantities_written_correctly(tmp_path):
    source = _build_template_workbook(tmp_path, "source5.xlsx")
    output_path = tmp_path / "output5.xlsx"

    generate_order_workbook(
        order_title="Test",
        order_lines=[OrderLine(row=3, quantity=100, free_quantity=20)],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    assert worksheet["D3"].value == 20


def test_absent_optional_notes_leave_optional_cells_blank(tmp_path):
    source = _build_template_workbook(tmp_path, "optional_blank_source.xlsx")
    output_path = tmp_path / "optional_blank_output.xlsx"

    generate_order_workbook(
        order_title="Customer without optional details",
        order_lines=[OrderLine(row=3, quantity=1, free_quantity=0, notes=None)],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    workbook, worksheet = _load(output_path)
    assert worksheet["F3"].value is None
    assert worksheet["F4"].value is None
    workbook.close()


def test_old_d_and_e_values_are_cleared_before_writing(tmp_path):
    source = _build_template_workbook(tmp_path, "source6.xlsx")
    # Simulate a previous order already sitting in the template copy.
    workbook = openpyxl.load_workbook(str(source))
    worksheet = workbook["Sheet1"]
    worksheet["D4"] = 999
    worksheet["E4"] = 50
    worksheet["H4"] = "old note"
    workbook.save(str(source))
    workbook.close()

    output_path = tmp_path / "output6.xlsx"
    generate_order_workbook(
        order_title="Test",
        order_lines=[OrderLine(row=3, quantity=5, free_quantity=0)],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    assert worksheet["C3"].value == 5
    assert worksheet["D3"].value == 0
    assert worksheet["F3"].value is None
    assert worksheet.max_row == 4


def test_remaining_value_and_total_formulas_are_rebuilt_for_compact_layout(tmp_path):
    source = _build_template_workbook(tmp_path, "source7.xlsx")
    output_path = tmp_path / "output7.xlsx"

    generate_order_workbook(
        order_title="Test",
        order_lines=[OrderLine(row=3, quantity=5, free_quantity=0)],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    assert worksheet["E3"].value == "=C3*B3"
    assert worksheet["E4"].value == "=SUM(E3:E3)"
    assert all(
        "#REF!" not in str(cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )


def test_prices_remain_unchanged(tmp_path):
    source = _build_template_workbook(tmp_path, "source8.xlsx")
    output_path = tmp_path / "output8.xlsx"

    generate_order_workbook(
        order_title="Test",
        order_lines=[OrderLine(row=3, quantity=5, free_quantity=0)],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    assert worksheet["B3"].value == 1200
    assert worksheet.max_column == 6


def test_styles_and_merged_cells_are_preserved(tmp_path):
    source = _build_template_workbook(tmp_path, "source9.xlsx")
    output_path = tmp_path / "output9.xlsx"

    generate_order_workbook(
        order_title="New Customer Name",
        order_lines=[],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    assert "A1:F1" in [str(mc) for mc in worksheet.merged_cells.ranges]
    assert worksheet["A1"].font.bold is True
    assert worksheet["A1"].font.color.rgb == "FFFF0000"
    assert worksheet["A1"].alignment.horizontal == "center"
    assert worksheet.column_dimensions["A"].width == 43.58203125 or round(worksheet.column_dimensions["A"].width, 2) == 43.58
    assert worksheet["B2"].fill.fgColor.rgb == "FFBDD7EE"
    assert worksheet["B2"].border.bottom.style == "thin"
    assert worksheet.row_dimensions[1].height == 24
    assert worksheet.row_dimensions[2].height == 30
    assert worksheet.row_dimensions[3].height == 22
    assert worksheet.print_area == "'Sheet1'!$A$1:$F$3"
    assert worksheet.page_setup.orientation == "portrait"


def test_totals_row_remains_intact(tmp_path):
    source = _build_template_workbook(tmp_path, "source10.xlsx")
    output_path = tmp_path / "output10.xlsx"

    generate_order_workbook(
        order_title="Test",
        order_lines=[OrderLine(row=3, quantity=5, free_quantity=0)],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    assert worksheet.max_row == 4
    assert worksheet.max_column == 6
    assert worksheet["E4"].value == "=SUM(E3:E3)"
    assert worksheet["E4"].font.bold is True


@pytest.mark.parametrize(
    ("price_type", "expected_header", "expected_prices"),
    [
        ("drug_store", "Price (Drug Store)", [1000, 500]),
        ("pharmacy", "Price (Pharmacy & Hospitals)", [1200, 600]),
    ],
)
def test_generated_sheet_physically_keeps_only_selected_pricing_branch(
    tmp_path, price_type, expected_header, expected_prices
):
    source = _build_template_workbook(tmp_path, f"{price_type}_source.xlsx")
    output_path = tmp_path / f"{price_type}_output.xlsx"

    # Rows 3 and 5 are intentionally non-adjacent in the catalog.
    generate_order_workbook(
        order_title="Compact order",
        order_lines=[
            OrderLine(row=3, quantity=2, free_quantity=1),
            OrderLine(row=5, quantity=4, free_quantity=0, notes="Urgent"),
        ],
        selected_price_type=price_type,
        output_path=output_path,
        source_path=source,
    )

    workbook, worksheet = _load(output_path)
    assert worksheet.max_row == 5
    assert worksheet.max_column == 6
    assert [worksheet.cell(2, column).value for column in range(1, 7)] == [
        "Product name",
        expected_header,
        "Qty",
        "Free",
        "Value",
        "Note",
    ]
    assert [worksheet[f"A{row}"].value for row in (3, 4)] == [
        "Alpha Tablet 50MG",
        "Gamma Injection",
    ]
    assert [worksheet[f"B{row}"].value for row in (3, 4)] == expected_prices
    assert [worksheet[f"C{row}"].value for row in (3, 4)] == [2, 4]
    assert [worksheet[f"E{row}"].value for row in (3, 4)] == [
        "=C3*B3",
        "=C4*B4",
    ]
    assert worksheet["E5"].value == "=SUM(E3:E4)"
    assert worksheet["F4"].value == "Urgent"
    assert not any(
        dimension.hidden for dimension in worksheet.column_dimensions.values()
    )
    assert all(
        "#REF!" not in str(cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )
    source_workbook, source_worksheet = _load(source)
    selected_price_column = "B" if price_type == "drug_store" else "C"
    selected_value_column = "F" if price_type == "drug_store" else "G"
    assert worksheet["B3"]._style == source_worksheet[f"{selected_price_column}3"]._style
    assert worksheet["E3"]._style == source_worksheet[f"{selected_value_column}3"]._style
    assert worksheet["E5"]._style == source_worksheet[f"{selected_value_column}15"]._style
    expected_width_columns = (
        ("A", "B", "D", "E", "F", "H")
        if price_type == "drug_store"
        else ("A", "C", "D", "E", "G", "H")
    )
    assert [worksheet.column_dimensions[letter].width for letter in "ABCDEF"] == [
        source_worksheet.column_dimensions[letter].width
        for letter in expected_width_columns
    ]
    source_workbook.close()
    workbook.close()

    # An XLSX is a ZIP package; a successful integrity read catches broken output
    # packages that Excel would otherwise attempt to repair.
    with zipfile.ZipFile(output_path) as package:
        assert package.testzip() is None


def test_one_product_order_has_no_unused_rows_or_columns(tmp_path):
    source = _build_template_workbook(tmp_path, "single_source.xlsx")
    output_path = tmp_path / "single_output.xlsx"

    generate_order_workbook(
        order_title="Single product",
        order_lines=[OrderLine(row=4, quantity=7, free_quantity=2)],
        selected_price_type="drug_store",
        output_path=output_path,
        source_path=source,
    )

    workbook, worksheet = _load(output_path)
    assert worksheet.calculate_dimension() == "A1:F4"
    assert worksheet["A3"].value == "Beta Syrup"
    assert worksheet["B3"].value == 2000
    assert worksheet["E3"].value == "=C3*B3"
    assert worksheet["E4"].value == "=SUM(E3:E3)"
    workbook.close()


# --- totals calculation -------------------------------------------------------------------


def test_selected_pharmacy_total_is_calculated_correctly(tmp_path):
    source = _build_template_workbook(tmp_path, "source11.xlsx")
    output_path = tmp_path / "output11.xlsx"

    result = generate_order_workbook(
        order_title="Test",
        order_lines=[
            OrderLine(row=3, quantity=2, free_quantity=0),  # Alpha, pharmacy price 1200
            OrderLine(row=4, quantity=3, free_quantity=0),  # Beta, pharmacy price 2400
        ],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    assert result.selected_order_total == (2 * 1200) + (3 * 2400)


def test_selected_drug_store_total_is_calculated_correctly(tmp_path):
    source = _build_template_workbook(tmp_path, "source12.xlsx")
    output_path = tmp_path / "output12.xlsx"

    result = generate_order_workbook(
        order_title="Test",
        order_lines=[
            OrderLine(row=3, quantity=2, free_quantity=0),  # Alpha, drug-store price 1000
            OrderLine(row=4, quantity=3, free_quantity=0),  # Beta, drug-store price 2000
        ],
        selected_price_type="drug_store",
        output_path=output_path,
        source_path=source,
    )

    assert result.selected_order_total == (2 * 1000) + (3 * 2000)


def test_free_quantity_does_not_affect_the_total(tmp_path):
    source = _build_template_workbook(tmp_path, "source13.xlsx")
    output_path = tmp_path / "output13.xlsx"

    result = generate_order_workbook(
        order_title="Test",
        order_lines=[OrderLine(row=3, quantity=2, free_quantity=1000)],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    assert result.selected_order_total == 2 * 1200


# --- filename sanitization -----------------------------------------------------------------


def test_output_filename_is_sanitized_for_unsafe_characters():
    unsafe = 'Bad/Name\\With:Illegal*Chars?"<>|Title'
    safe = sanitize_filename_component(unsafe)
    for char in '/\\:*?"<>|':
        assert char not in safe


def test_output_filename_sanitizer_matches_expected_pattern():
    safe = sanitize_filename_component("صيدلية العين - النجف")
    assert safe == "صيدلية_العين_النجف"


def test_output_filename_sanitizer_truncates_excessive_length():
    safe = sanitize_filename_component("A" * 500)
    assert len(safe) <= 80


def test_build_output_filename_produces_expected_shape():
    filename = build_output_filename("صيدلية العين - النجف", "2026-07-25")
    assert filename == "صيدلية_العين_النجف_2026-07-25.xlsx"


def test_build_output_filename_adds_readable_collision_suffix():
    filename = build_output_filename("صيدلية النخبة", "2026-07-25", 2)
    assert filename == "صيدلية_النخبة_2026-07-25_(2).xlsx"


# --- formula-injection protection -----------------------------------------------------------


@pytest.mark.parametrize("dangerous_title", ["=1+1", "+CMD|'/C calc'!A1", "-2+3", "@SUM(1,1)"])
def test_formula_injection_in_customer_title_is_neutralized(tmp_path, dangerous_title):
    source = _build_template_workbook(tmp_path, "source14.xlsx")
    output_path = tmp_path / "output14.xlsx"

    generate_order_workbook(
        order_title=dangerous_title,
        order_lines=[],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    cell = worksheet["A1"]
    assert cell.data_type != "f"
    assert cell.value.startswith("'")
    assert cell.value[1:] == dangerous_title


def test_formula_injection_in_product_notes_is_neutralized(tmp_path):
    source = _build_template_workbook(tmp_path, "source15.xlsx")
    output_path = tmp_path / "output15.xlsx"

    generate_order_workbook(
        order_title="Test",
        order_lines=[OrderLine(row=3, quantity=5, free_quantity=0, notes="=HYPERLINK(\"http://evil\")")],
        selected_price_type="pharmacy",
        output_path=output_path,
        source_path=source,
    )

    _, worksheet = _load(output_path)
    cell = worksheet["F3"]
    assert cell.data_type != "f"
    assert cell.value.startswith("'")
