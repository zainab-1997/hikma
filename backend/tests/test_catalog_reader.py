"""Tests for the read-only Excel catalog reader.

All tests build small, synthetic temporary workbooks (via pytest's tmp_path) that mirror
the verified Hikma template structure — none of them touch the real
backend/templates/Hikma orders.xlsx file.
"""

import hashlib
import os

import openpyxl
import pytest

from excel import catalog_reader
from excel.catalog_reader import (
    FIRST_PRODUCT_ROW,
    HEADER_ROW,
    LAST_PRODUCT_ROW,
    CatalogUnavailableError,
    find_duplicate_official_names,
    get_catalog_products,
)


def _build_workbook(tmp_path, filename, rows, sheet_name="Sheet1", row15_value=None):
    """Build a workbook mirroring the verified structure: title row 1, header row 2,
    product rows 3-14 (only the given `rows` populated, everything else left blank),
    and an optional row 15 totals-style value."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet["A1"] = "اسم الزبون"
    worksheet["A2"] = "Product name"
    for row, name in rows.items():
        worksheet.cell(row=row, column=1, value=name)
    if row15_value is not None:
        worksheet.cell(row=15, column=1, value=row15_value)
        worksheet.cell(row=15, column=6, value=f"=SUM(F{FIRST_PRODUCT_ROW}:F{LAST_PRODUCT_ROW})")
    path = tmp_path / filename
    workbook.save(str(path))
    return path


def _file_hash(path) -> str:
    with open(path, "rb") as handle:
        return hashlib.sha256(handle.read()).hexdigest()


@pytest.fixture(autouse=True)
def _clear_cache():
    catalog_reader.clear_catalog_cache()
    yield
    catalog_reader.clear_catalog_cache()


# --- structure -----------------------------------------------------------------------


def test_header_row_is_immediately_before_first_product_row():
    assert FIRST_PRODUCT_ROW == HEADER_ROW + 1


def test_rows_3_through_14_are_detected_as_products(tmp_path):
    rows = {row: f"Product {row}" for row in range(FIRST_PRODUCT_ROW, LAST_PRODUCT_ROW + 1)}
    path = _build_workbook(tmp_path, "full.xlsx", rows)

    catalog = get_catalog_products(template_path=str(path))

    assert len(catalog) == 12
    assert [p.row for p in catalog] == list(range(FIRST_PRODUCT_ROW, LAST_PRODUCT_ROW + 1))


def test_row_15_is_excluded_as_totals(tmp_path):
    rows = {3: "Alpha Tablet 50MG"}
    path = _build_workbook(tmp_path, "totals.xlsx", rows, row15_value="Looks Like A Product")

    catalog = get_catalog_products(template_path=str(path))

    assert [p.official_name for p in catalog] == ["Alpha Tablet 50MG"]
    assert all(p.row != 15 for p in catalog)


def test_header_row_content_is_never_treated_as_a_product(tmp_path):
    path = _build_workbook(tmp_path, "header.xlsx", {3: "Alpha Tablet 50MG"})

    catalog = get_catalog_products(template_path=str(path))

    assert all(p.official_name != "Product name" for p in catalog)


def test_blank_rows_within_the_product_range_are_ignored(tmp_path):
    rows = {3: "Alpha Tablet 50MG", 4: "", 5: "Beta Syrup", 8: "Gamma Injection"}
    path = _build_workbook(tmp_path, "blanks.xlsx", rows)

    catalog = get_catalog_products(template_path=str(path))

    assert [p.official_name for p in catalog] == ["Alpha Tablet 50MG", "Beta Syrup", "Gamma Injection"]


def test_correct_product_name_column_is_used(tmp_path):
    path = _build_workbook(tmp_path, "column.xlsx", {3: "Alpha Tablet 50MG"})
    workbook = openpyxl.load_workbook(str(path))
    workbook["Sheet1"].cell(row=3, column=2, value="This is not the product name")
    workbook.save(str(path))

    catalog = get_catalog_products(template_path=str(path))

    assert catalog[0].official_name == "Alpha Tablet 50MG"


def test_invalid_worksheet_is_rejected(tmp_path):
    path = _build_workbook(tmp_path, "wrong_sheet.xlsx", {3: "Alpha Tablet 50MG"}, sheet_name="Data")

    with pytest.raises(CatalogUnavailableError):
        get_catalog_products(template_path=str(path))


def test_missing_workbook_is_handled_safely():
    with pytest.raises(CatalogUnavailableError):
        get_catalog_products(template_path="/tmp/does-not-exist-hikma-catalog.xlsx")


# --- caching ---------------------------------------------------------------------------


def test_cache_reuses_result_when_file_unchanged(tmp_path, monkeypatch):
    path = _build_workbook(tmp_path, "cache_reuse.xlsx", {3: "Alpha Tablet 50MG"})

    call_count = {"n": 0}
    original_read = catalog_reader._read_catalog_products

    def spy(template_path):
        call_count["n"] += 1
        return original_read(template_path)

    monkeypatch.setattr(catalog_reader, "_read_catalog_products", spy)

    get_catalog_products(template_path=str(path))
    get_catalog_products(template_path=str(path))

    assert call_count["n"] == 1


def test_cache_refreshes_after_file_modification(tmp_path):
    path = _build_workbook(tmp_path, "cache_refresh.xlsx", {3: "Alpha Tablet 50MG"})
    first = get_catalog_products(template_path=str(path))
    assert [p.official_name for p in first] == ["Alpha Tablet 50MG"]

    _build_workbook(tmp_path, "cache_refresh.xlsx", {3: "Beta Syrup"})
    new_mtime = os.path.getmtime(str(path)) + 5
    os.utime(str(path), (new_mtime, new_mtime))

    second = get_catalog_products(template_path=str(path))
    assert [p.official_name for p in second] == ["Beta Syrup"]


# --- safety ----------------------------------------------------------------------------


def test_reading_the_catalog_never_modifies_the_source_file(tmp_path):
    rows = {row: f"Product {row}" for row in range(FIRST_PRODUCT_ROW, LAST_PRODUCT_ROW + 1)}
    path = _build_workbook(tmp_path, "unmodified.xlsx", rows, row15_value="Total")

    before = _file_hash(path)
    get_catalog_products(template_path=str(path))
    get_catalog_products(template_path=str(path))
    after = _file_hash(path)

    assert before == after


def test_workbook_is_opened_read_only(tmp_path, monkeypatch):
    path = _build_workbook(tmp_path, "readonly.xlsx", {3: "Alpha Tablet 50MG"})

    captured_kwargs = {}
    original_load = openpyxl.load_workbook

    def spy_load(*args, **kwargs):
        captured_kwargs.update(kwargs)
        return original_load(*args, **kwargs)

    monkeypatch.setattr(catalog_reader.openpyxl, "load_workbook", spy_load)

    get_catalog_products(template_path=str(path))

    assert captured_kwargs.get("read_only") is True


# --- duplicate detection -----------------------------------------------------------------


def test_duplicate_official_names_are_flagged(tmp_path):
    rows = {3: "Alpha Tablet 50MG", 4: "Beta Syrup", 5: "alpha tablet 50mg"}
    path = _build_workbook(tmp_path, "duplicates.xlsx", rows)

    catalog = get_catalog_products(template_path=str(path))
    duplicates = find_duplicate_official_names(catalog)

    assert duplicates == ["Alpha Tablet 50MG", "alpha tablet 50mg"]


def test_no_duplicates_when_all_names_are_unique(tmp_path):
    rows = {3: "Alpha Tablet 50MG", 4: "Beta Syrup"}
    path = _build_workbook(tmp_path, "no_duplicates.xlsx", rows)

    catalog = get_catalog_products(template_path=str(path))

    assert find_duplicate_official_names(catalog) == []
