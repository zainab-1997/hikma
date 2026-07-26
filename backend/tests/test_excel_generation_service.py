"""Tests for the confirmed-order validation and orchestration layer
(services/excel_generation_service.py). Catalog and template are always synthetic /
injected — none of these tests depend on the real Hikma workbook.
"""

from datetime import date

import openpyxl
import pytest
from pydantic import ValidationError

from excel.catalog_reader import CatalogProduct
from excel.order_writer import ExcelGenerationError
from models.generate_order_models import ConfirmedMatchedProduct, GenerateOrderRequest
from services.excel_generation_service import generate_excel_order, resolve_generated_file_path

CATALOG = (
    CatalogProduct(row=3, official_name="Alpha Tablet 50MG"),
    CatalogProduct(row=4, official_name="Beta Syrup"),
)


def _build_template_workbook(tmp_path, filename="template.xlsx"):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = "اسم الزبون"

    for index, (name, price_b, price_c) in enumerate([("Alpha Tablet 50MG", 1000, 1200), ("Beta Syrup", 2000, 2400)]):
        row = 3 + index
        worksheet.cell(row=row, column=1, value=name)
        worksheet.cell(row=row, column=2, value=price_b)
        worksheet.cell(row=row, column=3, value=price_c)
        worksheet.cell(row=row, column=6, value=f"=D{row}*B{row}")
        worksheet.cell(row=row, column=7, value=f"=D{row}*C{row}")

    worksheet.cell(row=15, column=6, value="=SUM(F3:F14)")
    worksheet.cell(row=15, column=7, value="=SUM(G3:G14)")

    path = tmp_path / filename
    workbook.save(str(path))
    return path


def _product(**overrides):
    defaults = dict(
        written_product_name="Alpha",
        matched_row=3,
        matched_official_name="Alpha Tablet 50MG",
        quantity=5,
        free_quantity=0,
    )
    defaults.update(overrides)
    return ConfirmedMatchedProduct(**defaults)


def _request(**overrides):
    defaults = dict(
        order_title="Test Customer",
        selected_price_type="pharmacy",
        products=[_product()],
        required_confirmations_resolved=True,
        order_notes=[],
        customer_name="Test Customer",
    )
    defaults.update(overrides)
    return GenerateOrderRequest(**defaults)


# --- model-level validation (structural, independent of catalog) --------------------------


def test_unresolved_product_cannot_be_constructed():
    with pytest.raises(ValidationError):
        ConfirmedMatchedProduct(written_product_name="Alpha", quantity=5)


def test_decimal_quantity_is_rejected():
    with pytest.raises(ValidationError):
        ConfirmedMatchedProduct(
            written_product_name="Alpha", matched_row=3, matched_official_name="Alpha Tablet 50MG", quantity=5.5
        )


def test_negative_quantity_is_rejected():
    with pytest.raises(ValidationError):
        ConfirmedMatchedProduct(
            written_product_name="Alpha", matched_row=3, matched_official_name="Alpha Tablet 50MG", quantity=-1
        )


def test_zero_quantity_is_rejected():
    with pytest.raises(ValidationError):
        ConfirmedMatchedProduct(
            written_product_name="Alpha", matched_row=3, matched_official_name="Alpha Tablet 50MG", quantity=0
        )


def test_negative_free_quantity_is_rejected():
    with pytest.raises(ValidationError):
        _product(free_quantity=-1)


def test_unknown_price_type_is_rejected():
    with pytest.raises(ValidationError):
        _request(selected_price_type="unknown")


# --- catalog revalidation ------------------------------------------------------------------


def test_invalid_product_row_is_rejected(tmp_path):
    request = _request(products=[_product(matched_row=99, matched_official_name="Nonexistent")])
    with pytest.raises(ExcelGenerationError):
        generate_excel_order(
            request,
            catalog=CATALOG,
            source_path=_build_template_workbook(tmp_path),
            output_dir=tmp_path / "out",
        )


def test_official_name_mismatch_is_rejected(tmp_path):
    request = _request(products=[_product(matched_row=3, matched_official_name="Wrong Name Entirely")])
    with pytest.raises(ExcelGenerationError):
        generate_excel_order(
            request,
            catalog=CATALOG,
            source_path=_build_template_workbook(tmp_path),
            output_dir=tmp_path / "out",
        )


def test_duplicate_matched_row_is_rejected(tmp_path):
    request = _request(
        products=[
            _product(matched_row=3, matched_official_name="Alpha Tablet 50MG", quantity=5),
            _product(matched_row=3, matched_official_name="Alpha Tablet 50MG", quantity=2),
        ]
    )
    with pytest.raises(ExcelGenerationError):
        generate_excel_order(
            request,
            catalog=CATALOG,
            source_path=_build_template_workbook(tmp_path),
            output_dir=tmp_path / "out",
        )


def test_unresolved_business_confirmations_block_generation(tmp_path):
    request = _request(required_confirmations_resolved=False)
    with pytest.raises(ExcelGenerationError):
        generate_excel_order(
            request,
            catalog=CATALOG,
            source_path=_build_template_workbook(tmp_path),
            output_dir=tmp_path / "out",
        )


# --- successful generation -----------------------------------------------------------------


def test_successful_generation_returns_expected_metadata(tmp_path):
    request = _request(products=[_product(matched_row=3, matched_official_name="Alpha Tablet 50MG", quantity=2)])
    response = generate_excel_order(
        request,
        catalog=CATALOG,
        source_path=_build_template_workbook(tmp_path),
        output_dir=tmp_path / "out",
    )

    assert response.filename.endswith(".xlsx")
    assert response.download_url == f"/api/orders/download/{response.filename}"
    assert response.selected_price_type == "pharmacy"
    assert response.selected_order_total == 2 * 1200
    assert (tmp_path / "out" / response.filename).exists()


def test_generated_filename_is_readable_customer_name_and_date(tmp_path):
    response = generate_excel_order(
        _request(order_title="Internal order title", customer_name="صيدلية النخبة"),
        catalog=CATALOG,
        source_path=_build_template_workbook(tmp_path),
        output_dir=tmp_path / "out",
        filename_date=date(2026, 7, 26),
    )

    assert response.filename == "صيدلية_النخبة_2026-07-26.xlsx"


def test_filename_collisions_use_incrementing_suffixes(tmp_path):
    output_dir = tmp_path / "out"
    source = _build_template_workbook(tmp_path)
    request = _request(customer_name="صيدلية النخبة")

    first = generate_excel_order(
        request, catalog=CATALOG, source_path=source, output_dir=output_dir,
        filename_date=date(2026, 7, 26),
    )
    second = generate_excel_order(
        request, catalog=CATALOG, source_path=source, output_dir=output_dir,
        filename_date=date(2026, 7, 26),
    )
    third = generate_excel_order(
        request, catalog=CATALOG, source_path=source, output_dir=output_dir,
        filename_date=date(2026, 7, 26),
    )

    assert [first.filename, second.filename, third.filename] == [
        "صيدلية_النخبة_2026-07-26.xlsx",
        "صيدلية_النخبة_2026-07-26_(2).xlsx",
        "صيدلية_النخبة_2026-07-26_(3).xlsx",
    ]
    assert len(list(output_dir.glob("*.xlsx"))) == 3


def test_order_notes_are_reported_as_excluded(tmp_path):
    request = _request(order_notes=["urgent"])
    response = generate_excel_order(
        request,
        catalog=CATALOG,
        source_path=_build_template_workbook(tmp_path),
        output_dir=tmp_path / "out",
    )
    assert response.excluded_order_notes is True


def test_no_order_notes_means_not_excluded(tmp_path):
    request = _request(order_notes=[])
    response = generate_excel_order(
        request,
        catalog=CATALOG,
        source_path=_build_template_workbook(tmp_path),
        output_dir=tmp_path / "out",
    )
    assert response.excluded_order_notes is False


# --- download identifier safety -------------------------------------------------------------


def test_resolve_generated_file_path_accepts_a_safe_filename(tmp_path):
    (tmp_path / "order.xlsx").write_bytes(b"fake xlsx content")
    resolved = resolve_generated_file_path("order.xlsx", base_dir=tmp_path)
    assert resolved == (tmp_path / "order.xlsx").resolve()


@pytest.mark.parametrize(
    "file_id",
    ["../secret.xlsx", "..", "a/../../b.xlsx", "/etc/passwd", "..\\secret.xlsx", "no-extension", ""],
)
def test_resolve_generated_file_path_blocks_traversal_and_invalid_ids(tmp_path, file_id):
    with pytest.raises(ExcelGenerationError):
        resolve_generated_file_path(file_id, base_dir=tmp_path)
