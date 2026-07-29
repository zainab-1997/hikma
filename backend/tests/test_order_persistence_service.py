"""Tests for services/order_persistence_service.py — the orchestration layer that ties
Excel generation to database persistence.

Every test uses a synthetic temporary template and a temporary SQLite database — none of
them touch backend/templates/Hikma orders.xlsx or backend/database/app.db.
"""

import hashlib
from datetime import date
from unittest.mock import patch

import openpyxl
import pytest
from sqlalchemy import func, select

from database.models import Order
from database.session import init_db, session_scope
from excel.catalog_reader import CatalogProduct
from excel.order_writer import ExcelGenerationError
from models.generate_order_models import ConfirmedMatchedProduct, GenerateOrderRequest
from services.order_persistence_service import generate_and_persist_order, get_order_detail

CATALOG = (
    CatalogProduct(row=3, official_name="Alpha Tablet 50MG"),
    CatalogProduct(row=4, official_name="Beta Syrup"),
)


def _build_template_workbook(tmp_path, filename="template.xlsx"):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = "اسم الزبون"

    for index, (name, price_b, price_c) in enumerate([("Alpha Tablet 50MG", 1000, 1200), ("Beta Syrup", 2000, 2400)]):
        row = 3 + index
        worksheet.cell(row=row, column=1, value=name)
        worksheet.cell(row=row, column=2, value=price_b)
        worksheet.cell(row=row, column=3, value=price_c)
        worksheet.cell(row=row, column=6, value=f"=D{row}*B{row}")
        worksheet.cell(row=row, column=7, value=f"=D{row}*C{row}")

    worksheet.cell(row=15, column=6, value="=SUM(F3:F14)")
    worksheet.cell(row=15, column=7, value="=SUM(G3:G14)")

    path = tmp_path / filename
    workbook.save(str(path))
    return path


def _file_hash(path) -> str:
    with open(path, "rb") as handle:
        return hashlib.sha256(handle.read()).hexdigest()


def _db_url(tmp_path) -> str:
    return f"sqlite:///{tmp_path}/test.db"


def _product(**overrides):
    defaults = dict(
        written_product_name="Alpha",
        matched_row=3,
        matched_official_name="Alpha Tablet 50MG",
        quantity=5,
        free_quantity=0,
    )
    defaults.update(overrides)
    return ConfirmedMatchedProduct(**defaults)


def _request(**overrides):
    defaults = dict(
        order_title="Test Customer",
        selected_price_type="pharmacy",
        products=[_product()],
        required_confirmations_resolved=True,
        order_notes=[],
        customer_name="Test Customer",
        customer_type="pharmacy",
        governorate="النجف",
    )
    defaults.update(overrides)
    return GenerateOrderRequest(**defaults)


def _setup(tmp_path):
    url = _db_url(tmp_path)
    init_db(database_url=url)
    source = _build_template_workbook(tmp_path)
    output_dir = tmp_path / "generated"
    return url, source, output_dir


def test_transit_route_title_is_identical_in_persistence_history_and_workbook(tmp_path):
    url, source, output_dir = _setup(tmp_path)
    response = generate_and_persist_order(
        _request(
            order_title="stale - Transit - title",
            customer_name="مذخر ساوا",
            is_transit=True,
            primary_customer="مذخر ساوا",
            destination_customer="مستشفى الكوثر",
        ),
        database_url=url,
        catalog=CATALOG,
        source_path=source,
        output_dir=output_dir,
    )

    expected = "مذخر ساوا - ترانزيت - مستشفى الكوثر - النجف"
    detail = get_order_detail(response.order_id, database_url=url)
    workbook = openpyxl.load_workbook(output_dir / response.filename, data_only=False)

    assert detail is not None
    assert detail.order_title == expected
    assert workbook.active["A1"].value == expected
    assert "ترانزيت" in response.filename
    assert "Transit" not in response.filename


# --- no save when generation fails ----------------------------------------------------------


def test_no_order_saved_when_excel_generation_fails(tmp_path):
    url, source, output_dir = _setup(tmp_path)
    request = _request(products=[_product(matched_row=99, matched_official_name="Nonexistent")])

    with pytest.raises(ExcelGenerationError):
        generate_and_persist_order(
            request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
        )

    with session_scope(url) as session:
        assert session.execute(select(func.count()).select_from(Order)).scalar_one() == 0


# --- cleanup on persistence failure (strategy A) ------------------------------------------


def test_generated_file_deleted_when_persistence_fails(tmp_path):
    url, source, output_dir = _setup(tmp_path)
    request = _request()

    with patch(
        "services.order_persistence_service.order_repository.create_order_with_products",
        side_effect=RuntimeError("simulated database failure"),
    ):
        with pytest.raises(ExcelGenerationError):
            generate_and_persist_order(
                request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
            )

    # The Excel file was generated, then removed because the order couldn't be saved.
    assert list(output_dir.glob("*.xlsx")) == []

    with session_scope(url) as session:
        assert session.execute(select(func.count()).select_from(Order)).scalar_one() == 0


# --- successful persistence ------------------------------------------------------------------


def test_successful_generation_persists_order_and_returns_metadata(tmp_path):
    url, source, output_dir = _setup(tmp_path)
    request = _request(products=[_product(matched_row=3, matched_official_name="Alpha Tablet 50MG", quantity=2)])

    response = generate_and_persist_order(
        request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )

    assert response.order_id
    assert response.order_number.startswith("HIK-")
    assert response.selected_order_total == 2 * 1200
    assert (output_dir / response.filename).exists()

    with session_scope(url) as session:
        assert session.execute(select(func.count()).select_from(Order)).scalar_one() == 1


def test_stored_total_comes_from_generation_result_not_request(tmp_path):
    # GenerateOrderRequest has no "total" field at all — this proves the persisted total
    # is derived purely from catalog price * quantity, computed server-side.
    url, source, output_dir = _setup(tmp_path)
    request = _request(products=[_product(matched_row=4, matched_official_name="Beta Syrup", quantity=3)])

    response = generate_and_persist_order(
        request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )

    assert response.selected_order_total == 3 * 2400  # Beta Syrup pharmacy price

    with session_scope(url) as session:
        order = session.get(Order, response.order_id)
        assert order.selected_order_total == 3 * 2400


# --- idempotency -----------------------------------------------------------------------------


def test_client_request_id_idempotency_returns_existing_order_without_regenerating(tmp_path):
    url, source, output_dir = _setup(tmp_path)
    request = _request(client_request_id="req-xyz-1")

    first = generate_and_persist_order(
        request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )

    with patch("services.order_persistence_service.generate_excel_order") as mock_generate:
        second = generate_and_persist_order(
            request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
        )
        mock_generate.assert_not_called()

    assert second.order_id == first.order_id
    assert second.order_number == first.order_number

    with session_scope(url) as session:
        assert session.execute(select(func.count()).select_from(Order)).scalar_one() == 1


def test_browser_double_click_with_same_request_id_creates_only_one_order(tmp_path):
    url, source, output_dir = _setup(tmp_path)
    request = _request(client_request_id="double-click-1")

    first = generate_and_persist_order(
        request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )
    second = generate_and_persist_order(
        request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )

    assert second.order_id == first.order_id
    assert second.filename == first.filename
    assert len(list(output_dir.glob("*.xlsx"))) == 1


def test_network_retry_with_same_request_id_returns_original_response(tmp_path):
    url, source, output_dir = _setup(tmp_path)
    request = _request(client_request_id="network-retry-1")

    first = generate_and_persist_order(
        request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )
    retry = generate_and_persist_order(
        request, database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )

    assert retry.model_dump() == first.model_dump()


def test_different_client_request_ids_create_separate_orders(tmp_path):
    url, source, output_dir = _setup(tmp_path)

    first = generate_and_persist_order(
        _request(client_request_id="req-a"), database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )
    second = generate_and_persist_order(
        _request(client_request_id="req-b"), database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )

    assert first.order_id != second.order_id
    with session_scope(url) as session:
        assert session.execute(select(func.count()).select_from(Order)).scalar_one() == 2


def test_same_customer_and_products_after_two_days_are_independent_orders(tmp_path):
    url, source, output_dir = _setup(tmp_path)

    monday = generate_and_persist_order(
        _request(customer_name="صيدلية النخبة", order_title="صيدلية النخبة", client_request_id="monday"),
        database_url=url,
        catalog=CATALOG,
        source_path=source,
        output_dir=output_dir,
        filename_date=date(2026, 7, 20),
    )
    wednesday = generate_and_persist_order(
        _request(customer_name="صيدلية النخبة", order_title="صيدلية النخبة", client_request_id="wednesday"),
        database_url=url,
        catalog=CATALOG,
        source_path=source,
        output_dir=output_dir,
        filename_date=date(2026, 7, 22),
    )

    assert monday.order_id != wednesday.order_id
    assert monday.filename == "صيدلية_النخبة_2026-07-20.xlsx"
    assert wednesday.filename == "صيدلية_النخبة_2026-07-22.xlsx"
    with session_scope(url) as session:
        orders = session.execute(select(Order).order_by(Order.created_at)).scalars().all()
        assert len(orders) == 2
        assert [order.customer_name for order in orders] == ["صيدلية النخبة", "صيدلية النخبة"]
        assert [order.products[0].quantity for order in orders] == [5, 5]


def test_identical_contents_without_idempotency_key_always_create_new_orders(tmp_path):
    url, source, output_dir = _setup(tmp_path)

    first = generate_and_persist_order(
        _request(client_request_id=None), database_url=url, catalog=CATALOG,
        source_path=source, output_dir=output_dir, filename_date=date(2026, 7, 26),
    )
    second = generate_and_persist_order(
        _request(client_request_id=None), database_url=url, catalog=CATALOG,
        source_path=source, output_dir=output_dir, filename_date=date(2026, 7, 26),
    )

    assert first.order_id != second.order_id
    assert first.filename == "Test_Customer_2026-07-26.xlsx"
    assert second.filename == "Test_Customer_2026-07-26_(2).xlsx"


# --- source workbook safety, exercised through the full persistence flow --------------------


def test_source_template_never_modified_through_full_persistence_flow(tmp_path):
    url, source, output_dir = _setup(tmp_path)
    before_hash = _file_hash(source)

    generate_and_persist_order(
        _request(), database_url=url, catalog=CATALOG, source_path=source, output_dir=output_dir
    )

    assert _file_hash(source) == before_hash
