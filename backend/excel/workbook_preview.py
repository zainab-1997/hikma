"""Read-only projection of the exact generated workbook used by download and email."""

import hashlib
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from models.generate_order_models import (
    WorkbookPreview,
    WorkbookPreviewCell,
    WorkbookPreviewRow,
)

_MULTIPLY_FORMULA = re.compile(r"=([A-Z]+)(\d+)\*([A-Z]+)(\d+)$")
_SUM_FORMULA = re.compile(r"=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$", re.IGNORECASE)


def _rgb(color) -> str | None:
    if color is None or color.type != "rgb" or not color.rgb:
        return None
    return str(color.rgb)[-6:]


def _calculated_value(worksheet, cell):
    if not isinstance(cell.value, str) or not cell.value.startswith("="):
        return cell.value
    multiply = _MULTIPLY_FORMULA.fullmatch(cell.value)
    if multiply:
        left = worksheet[f"{multiply.group(1)}{multiply.group(2)}"].value or 0
        right = worksheet[f"{multiply.group(3)}{multiply.group(4)}"].value or 0
        return left * right
    summed = _SUM_FORMULA.fullmatch(cell.value)
    if summed and summed.group(1) == summed.group(3):
        return sum(
            _calculated_value(worksheet, worksheet.cell(row=row, column=cell.column)) or 0
            for row in range(int(summed.group(2)), int(summed.group(4)) + 1)
        )
    return None


def build_workbook_preview(path: Path) -> WorkbookPreview:
    """Open one read-only workbook object and project its populated range."""
    workbook_sha256 = hashlib.sha256(path.read_bytes()).hexdigest()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        worksheet = workbook.active
        merged_by_start = {}
        covered = set()
        for merged in worksheet.merged_cells.ranges:
            merged_by_start[(merged.min_row, merged.min_col)] = merged.max_col - merged.min_col + 1
            for row in range(merged.min_row, merged.max_row + 1):
                for column in range(merged.min_col, merged.max_col + 1):
                    if (row, column) != (merged.min_row, merged.min_col):
                        covered.add((row, column))

        rows = []
        for row_number in range(1, worksheet.max_row + 1):
            cells = []
            for column in range(1, worksheet.max_column + 1):
                if (row_number, column) in covered:
                    continue
                cell = worksheet.cell(row=row_number, column=column)
                value = _calculated_value(worksheet, cell)
                cells.append(
                    WorkbookPreviewCell(
                        column=column,
                        value=value,
                        formula=cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
                        colspan=merged_by_start.get((row_number, column), 1),
                        font_bold=bool(cell.font.bold),
                        font_color=_rgb(cell.font.color),
                        fill_color=_rgb(cell.fill.fgColor) if cell.fill.fill_type else None,
                        horizontal_alignment=cell.alignment.horizontal,
                        number_format=cell.number_format,
                        border_top=cell.border.top.style,
                        border_right=cell.border.right.style,
                        border_bottom=cell.border.bottom.style,
                        border_left=cell.border.left.style,
                    )
                )
            rows.append(
                WorkbookPreviewRow(
                    row=row_number,
                    height=worksheet.row_dimensions[row_number].height,
                    cells=cells,
                )
            )
        return WorkbookPreview(
            sheet_name=worksheet.title,
            rows=rows,
            column_widths=[
                worksheet.column_dimensions[get_column_letter(column)].width
                for column in range(1, worksheet.max_column + 1)
            ],
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
            workbook_sha256=workbook_sha256,
        )
    finally:
        workbook.close()
