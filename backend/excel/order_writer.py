"""Writes a confirmed order into a fresh copy of the company Excel template.

Safety model: the source template (backend/templates/Hikma orders.xlsx) is never opened
by openpyxl in this module. It is only ever duplicated at the filesystem level
(`shutil.copy2`), which is a byte-for-byte read-only copy — nothing that can corrupt or
alter the source. openpyxl only ever touches the *copy*, and only that copy is saved.

Source workbook structure (backend/templates/Hikma orders.xlsx, Sheet1):
- A1:H1 is a single merged cell holding the customer/order-title field (in the blank
  template it holds the static label "اسم الزبون" — there is no separate value cell, so
  this merged cell IS the designated field; its value is replaced, its formatting is not).
- Row 2 is the header row.
- Rows 3-14 are product rows: A=name, B=drug-store price, C=pharmacy/hospital price,
  D=Qty, E=FQty, F='=D*B', G='=D*C', H=Note.
- Row 15 is a totals row: F15='=SUM(F3:F14)', G15='=SUM(G3:G14)'.
- The sheet has no cell anywhere outside A1:H15 — there is no safe place for a general,
  order-level note, only per-product notes (column H).

The generated copy is physically compacted to six columns:
Product name, selected Price, Qty, Free, Value, Note. Unordered catalog rows and the
unused pricing/value branch are deleted; formulas are rebuilt against the compact layout.
"""

import re
import shutil
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from config.settings import get_settings

WORKSHEET_NAME = "Sheet1"
CUSTOMER_TITLE_CELL = "A1"
FIRST_PRODUCT_ROW = 3
LAST_PRODUCT_ROW = 14

PRODUCT_NAME_COLUMN = 1  # A
DRUG_STORE_PRICE_COLUMN = 2  # B
PHARMACY_PRICE_COLUMN = 3  # C
QTY_COLUMN = 4  # D
FREE_QTY_COLUMN = 5  # E
DRUG_STORE_VALUE_COLUMN = 6  # F, '=D*B'
PHARMACY_VALUE_COLUMN = 7  # G, '=D*C'
NOTE_COLUMN = 8  # H

FINAL_PRICE_COLUMN = 2  # B
FINAL_QTY_COLUMN = 3  # C
FINAL_FREE_QTY_COLUMN = 4  # D
FINAL_VALUE_COLUMN = 5  # E
FINAL_NOTE_COLUMN = 6  # F
FINAL_COLUMN_COUNT = 6

_FORMULA_INJECTION_PREFIXES = ("=", "+", "-", "@")
_FILENAME_UNSAFE_CHARS = re.compile(r'[\\/:*?"<>|\x00-\x1f]')
_WHITESPACE_RUN = re.compile(r"\s+")
_UNDERSCORE_RUN = re.compile(r"_+")


class ExcelGenerationError(Exception):
    """Raised when the confirmed order cannot be safely written into the template."""

    status_code = 422


class TemplateUnavailableError(ExcelGenerationError):
    """Raised when the source template itself can't be found or opened."""

    status_code = 503


@dataclass(frozen=True)
class OrderLine:
    row: int
    quantity: int
    free_quantity: int
    notes: str | None = None


@dataclass(frozen=True)
class WriteResult:
    output_path: Path
    selected_order_total: int


def neutralize_cell_text(value: str) -> str:
    """Prevent formula injection: in Excel, a cell value starting with =, +, -, or @ is
    interpreted as a formula. Prefixing with an apostrophe forces it to stay plain text."""
    if value and value[0] in _FORMULA_INJECTION_PREFIXES:
        return f"'{value}"
    return value


def sanitize_filename_component(value: str) -> str:
    value = unicodedata.normalize("NFKC", value)
    value = _FILENAME_UNSAFE_CHARS.sub("", value)
    value = value.replace("-", "_")
    value = _WHITESPACE_RUN.sub("_", value.strip())
    value = _UNDERSCORE_RUN.sub("_", value)
    value = value.strip("_.")
    return value[:80] or "order"


def build_output_filename(
    order_title: str, order_date: str, collision_number: int | None = None
) -> str:
    safe_title = sanitize_filename_component(order_title)
    suffix = f"_({collision_number})" if collision_number is not None else ""
    return f"{safe_title}_{order_date}{suffix}.xlsx"


def _compact_generated_worksheet(worksheet, order_lines: list[OrderLine], selected_price_type: str) -> int:
    """Physically remove unused product rows and the unselected pricing branch.

    Returns the final totals-row number. This operates only on the already-copied
    workbook loaded by generate_order_workbook().
    """
    ordered_rows = {line.row for line in order_lines}
    original_print_area = worksheet.print_area
    original_widths = {
        column: worksheet.column_dimensions[column].width
        for column in "ABCDEFGH"
    }
    original_row_heights = {
        row: worksheet.row_dimensions[row].height
        for row in range(1, LAST_PRODUCT_ROW + 2)
    }

    # Remove unordered catalog rows in reverse so original row numbers remain valid.
    for row in range(LAST_PRODUCT_ROW, FIRST_PRODUCT_ROW - 1, -1):
        if row not in ordered_rows:
            worksheet.delete_rows(row, 1)

    totals_row = FIRST_PRODUCT_ROW + len(order_lines)

    # The only source merge spans all eight original columns. Unmerge before physical
    # column deletion, then restore it across the six retained columns.
    for merged_range in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(merged_range))

    if selected_price_type == "drug_store":
        kept_original_columns = ("A", "B", "D", "E", "F", "H")
        worksheet.delete_cols(7, 1)  # Pharmacy/Hospital Value (G)
        worksheet.delete_cols(3, 1)  # Pharmacy/Hospital Price (C)
        price_header = "Price (Drug Store)"
    else:
        kept_original_columns = ("A", "C", "D", "E", "G", "H")
        worksheet.delete_cols(6, 1)  # Drug Store Value (F)
        worksheet.delete_cols(2, 1)  # Drug Store Price (B)
        price_header = "Price (Pharmacy & Hospitals)"

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=FINAL_COLUMN_COUNT)

    headers = ("Product name", price_header, "Qty", "Free", "Value", "Note")
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=2, column=column).value = header

    # Rebuild every formula after deletion. No translated source formula is trusted.
    for row in range(FIRST_PRODUCT_ROW, totals_row):
        worksheet.cell(row=row, column=FINAL_VALUE_COLUMN).value = (
            f"=C{row}*B{row}"
        )
    worksheet.cell(row=totals_row, column=FINAL_VALUE_COLUMN).value = (
        f"=SUM(E{FIRST_PRODUCT_ROW}:E{totals_row - 1})"
        if order_lines
        else "=0"
    )

    # Preserve widths for the six columns that survived, in their new positions.
    for final_index, original_column in enumerate(kept_original_columns, start=1):
        final_letter = get_column_letter(final_index)
        worksheet.column_dimensions[final_letter].width = original_widths[original_column]
        worksheet.column_dimensions[final_letter].hidden = False
    for column in ("G", "H"):
        if column in worksheet.column_dimensions:
            del worksheet.column_dimensions[column]

    # openpyxl moves cells but not all row-dimension metadata. Reapply the source
    # heights to the compact row sequence and discard dimensions below the total.
    final_source_rows = [1, 2, *sorted(ordered_rows), LAST_PRODUCT_ROW + 1]
    for final_row, source_row in enumerate(final_source_rows, start=1):
        worksheet.row_dimensions[final_row].height = original_row_heights[source_row]
        worksheet.row_dimensions[final_row].hidden = False
    for row in list(worksheet.row_dimensions):
        if row > totals_row:
            del worksheet.row_dimensions[row]

    if original_print_area:
        worksheet.print_area = f"A1:F{totals_row}"

    worksheet.sheet_view.showGridLines = False
    return totals_row


def generate_order_workbook(
    *,
    order_title: str,
    order_lines: list[OrderLine],
    selected_price_type: str,
    output_path: Path,
    source_path: Path | None = None,
) -> WriteResult:
    """Copy the template to output_path and fill in the confirmed order.

    Callers are responsible for validating order_lines' rows against the live catalog
    before calling this function — this function trusts its input and only concerns
    itself with safely writing it into the workbook.

    Pass source_path explicitly in tests to use a temporary workbook instead of the
    real, settings-configured Hikma template.
    """
    if source_path is None:
        source_path = Path(get_settings().excel_template_path)

    if not source_path.exists():
        raise TemplateUnavailableError("The order template could not be found.")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        shutil.copy2(source_path, output_path)
    except OSError as exc:
        raise ExcelGenerationError("The order template could not be copied.") from exc

    try:
        workbook = openpyxl.load_workbook(output_path)
    except Exception as exc:
        raise TemplateUnavailableError("The order template could not be opened.") from exc

    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise TemplateUnavailableError(
                f'The order template is missing the expected "{WORKSHEET_NAME}" worksheet.'
            )
        worksheet = workbook[WORKSHEET_NAME]

        worksheet[CUSTOMER_TITLE_CELL] = neutralize_cell_text(order_title)

        # Clear any previous order's entry values before writing the new one. Prices and
        # product names are never changed before the generated copy is compacted.
        for row in range(FIRST_PRODUCT_ROW, LAST_PRODUCT_ROW + 1):
            worksheet.cell(row=row, column=QTY_COLUMN).value = None
            worksheet.cell(row=row, column=FREE_QTY_COLUMN).value = None
            worksheet.cell(row=row, column=NOTE_COLUMN).value = None

        price_column = DRUG_STORE_PRICE_COLUMN if selected_price_type == "drug_store" else PHARMACY_PRICE_COLUMN

        selected_order_total = 0
        for line in order_lines:
            worksheet.cell(row=line.row, column=QTY_COLUMN).value = line.quantity
            worksheet.cell(row=line.row, column=FREE_QTY_COLUMN).value = line.free_quantity
            if line.notes:
                worksheet.cell(row=line.row, column=NOTE_COLUMN).value = neutralize_cell_text(line.notes)

            unit_price = worksheet.cell(row=line.row, column=price_column).value or 0
            selected_order_total += int(unit_price) * line.quantity

        _compact_generated_worksheet(worksheet, order_lines, selected_price_type)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output_path)
    finally:
        workbook.close()

    return WriteResult(output_path=output_path, selected_order_total=selected_order_total)
